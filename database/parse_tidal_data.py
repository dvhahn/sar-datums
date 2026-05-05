"""
parse_tidal_data.py
Parses the tidal current vector data from the Excel Data sheet
and inserts it into the PostgreSQL tidal_vectors table.

The Excel file stores data in a "wide" format:
  - Each row = one geographic point (lat, lon)
  - Columns = vx/vy pairs for each time step (0.0h to 12.5h)
  - Two sections: Neap and Spring (separated by an empty column)

This script normalises the data into rows of:
  (location, lat, lon, tide_type, time_step, vx, vy)

Usage:
    python database/parse_tidal_data.py path/to/excel_file.xlsm

Requirements:
    pip install openpyxl psycopg2-binary
"""

import sys
import time
import argparse
import openpyxl
from io import StringIO
from database.db_config import get_connection

# Excel Data sheet column mapping (1-indexed, openpyxl convention)
LAT_COL = 7        # Column G
LON_COL = 8        # Column H
NEAP_START_COL = 9  # Column I (Neap 0.0h vx)

DATA_START_ROW = 2  # First data row (after header)

# Time steps: 0.0, 0.1, 0.2 ... 12.5 = 126 steps
TIME_STEPS = [round(i * 0.1, 1) for i in range(126)]
COLS_PER_TIDE_TYPE = len(TIME_STEPS) * 2  # 126 steps x 2 (vx, vy) = 252 columns


def find_spring_start(ws):
    """Scan the header row to find where the Spring section begins.
    Looks for a 'lat' header after the Neap columns."""
    header_row = 1
    for col in range(NEAP_START_COL + COLS_PER_TIDE_TYPE, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip().lower() == 'lat':
            return col
    return None


def parse_and_insert(excel_path, target="local"):
    print(f"Opening {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Data']

    spring_lat_col = find_spring_start(ws)
    if spring_lat_col is None:
        print("ERROR: Could not find Spring section in Data sheet.")
        sys.exit(1)

    spring_start_col = spring_lat_col + 2
    print(f"Neap starts at col {NEAP_START_COL}, Spring starts at col {spring_start_col}")

    conn = get_connection(target=target)
    cur = conn.cursor()

    try:
        # Clear existing data
        print(f"Truncating table on {target}...")
        cur.execute("TRUNCATE TABLE tidal_vectors RESTART IDENTITY;")
        conn.commit()

        print("Parsing and inserting data...")

        # 4. OPTIMIZATION: Use larger batches for AWS to combat network latency
        batch_size = 1000 if target == "aws" else 500

        total_rows = 0
        total_db_rows = 0
        start_time = time.time()
        buffer = StringIO()

        for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
            lat = row[LAT_COL - 1]
            lon = row[LON_COL - 1]

            if lat is None or lon is None:
                continue

            lat, lon = float(lat), float(lon)

            # --- Neap data ---
            for i, ts in enumerate(TIME_STEPS):
                col_offset = NEAP_START_COL - 1 + (i * 2)
                vx = row[col_offset] if col_offset < len(row) else None
                vy = row[col_offset + 1] if col_offset + 1 < len(row) else None
                if vx is not None and vy is not None:
                    buffer.write(f"SRID=4326;POINT({lon} {lat})\t{lat}\t{lon}\tneap\t{ts}\t{vx}\t{vy}\n")
                    total_db_rows += 1

            # --- Spring data ---
            for i, ts in enumerate(TIME_STEPS):
                col_offset = spring_start_col - 1 + (i * 2)
                vx = row[col_offset] if col_offset < len(row) else None
                vy = row[col_offset + 1] if col_offset + 1 < len(row) else None
                if vx is not None and vy is not None:
                    buffer.write(f"SRID=4326;POINT({lon} {lat})\t{lat}\t{lon}\tspring\t{ts}\t{vx}\t{vy}\n")
                    total_db_rows += 1

            total_rows += 1

            if total_rows % batch_size == 0:
                buffer.seek(0)
                cur.copy_from(buffer, 'tidal_vectors',
                              columns=('location', 'lat', 'lon', 'tide_type', 'time_step', 'vx', 'vy'),
                              sep='\t')
                conn.commit()
                buffer = StringIO()

                elapsed = time.time() - start_time
                print(
                    f"  {total_rows:,} locations processed - {total_rows / (elapsed if elapsed > 0 else 1):.0f} rows/sec")

        # Final flush
        if buffer.tell() > 0:
            buffer.seek(0)
            cur.copy_from(buffer, 'tidal_vectors',
                          columns=('location', 'lat', 'lon', 'tide_type', 'time_step', 'vx', 'vy'),
                          sep='\t')
            conn.commit()

    finally:
        cur.close()
        conn.close()
        wb.close()

    elapsed = time.time() - start_time
    print(f"\nDone! Total {total_db_rows:,} vector rows inserted in {elapsed:.1f}s")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("path", help="Path to Excel file")
    parser.add_argument("--target", default="local", choices=["local", "aws"], help="DB to target")
    args = parser.parse_args()

    parse_and_insert(args.path, args.target)